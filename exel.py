import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

ARQUIVO_EXCEL = "comparativo_precos.xlsx"

def tratar_preco(valor):
    """Converte valor para float."""
    if not valor:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    valor_str = str(valor).strip()
    # Remove símbolos de moeda e espaços
    valor_str = valor_str.lower().replace('r$', '').replace('us$', '').strip()
    
    # Lógica para tratar milhar e decimal (padrão BR: 1.000,00)
    if ',' in valor_str and '.' in valor_str:
        valor_str = valor_str.replace('.', '') # Remove milhar
        valor_str = valor_str.replace(',', '.') # Virgula vira ponto
    elif ',' in valor_str:
        valor_str = valor_str.replace(',', '.')
    
    try:
        return float(valor_str)
    except ValueError:
        return 0.0

def formatar_planilha(ws, loja):
    """Aplica o css visual na aba."""
    # Cores e Fontes
    COR_TITULO_BG = "003366"
    COR_HEADER_BG = "1F497D"
    COR_ZEBRA_BG  = "F2F2F2"
    COR_TEXTO_BRANCO = "FFFFFF" 

    titulo_font = Font(name='Calibri', size=16, bold=True, color=COR_TEXTO_BRANCO)
    header_font = Font(name='Calibri', size=12, bold=True, color=COR_TEXTO_BRANCO)
    
    titulo_fill = PatternFill(start_color=COR_TITULO_BG, end_color=COR_TITULO_BG, fill_type="solid")
    header_fill = PatternFill(start_color=COR_HEADER_BG, end_color=COR_HEADER_BG, fill_type="solid")
    zebra_fill = PatternFill(start_color=COR_ZEBRA_BG, end_color=COR_ZEBRA_BG, fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color="BFBFBF"), right=Side(style='thin', color="BFBFBF"),
        top=Side(style='thin', color="BFBFBF"), bottom=Side(style='thin', color="BFBFBF")
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Css titulo
    ws.merge_cells('A1:C1')
    cell_title = ws['A1']
    cell_title.value = f"RELATÓRIO DE PREÇOS - {loja.upper()}"
    cell_title.font = titulo_font
    cell_title.fill = titulo_fill
    cell_title.alignment = align_center
    ws.row_dimensions[1].height = 35

    # css Cabeçalho
    headers = ["Nome do Produto", "Preço", "Link"]
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Css linhas 
    max_row = ws.max_row
    if max_row >= 3:
        # Define a área de dados
        data_range = ws[f'A3:C{max_row}']
        
        for row_idx, row in enumerate(data_range, start=3):
            for cell in row:
                cell.border = thin_border
                #Efeito zebra nas linhas -> simplismente por aparência
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
                
                # Alinhamentos específicos
                if cell.column == 1: # Nome
                    cell.alignment = align_left_wrap
                elif cell.column == 2: # Preço
                    cell.alignment = align_center
                    cell.font = Font(bold=True)
                elif cell.column == 3: # Link
                    cell.alignment = align_center

    # Css largura
    ws.auto_filter.ref = f"A2:C{max_row}"
    ws.freeze_panes = "A3"
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15

def salvar_dados(dados, loja):
    """
    Salva os dados de uma loja específica. 
    Se o arquivo já existe, adiciona ou atualiza a aba daquela loja.
    """
    if not dados:
        print(f"[{loja}] Nenhum dado para salvar.")
        return

    # Carrega o arquivo ou cria se ele ja existe
    if os.path.exists(ARQUIVO_EXCEL):

        #Linha responsável por carregar
        try:
            wb = load_workbook(ARQUIVO_EXCEL)
        except PermissionError:
            print(f"\n[ERRO] O arquivo '{ARQUIVO_EXCEL}' está aberto. Feche-o e tente novamente.")
            return
    else:
        #linha responsável por criar
        wb = Workbook()
        # Se acabou de criar, remove a aba padrão 'Sheet' se ela existir
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    # Se a aba da loja já existe, deleta ela para criar uma nova atualizada
    if loja in wb.sheetnames:
        del wb[loja]
    
    # Cria a nova aba
    ws = wb.create_sheet(title=loja)
    
    # Escreve os dados
    start_row = 3
    for i, linha in enumerate(dados):
        # Nome
        ws.cell(row=start_row + i, column=1, value=linha[0])
        
        # Preço
        preco_formatado = tratar_preco(linha[1])
        cell_preco = ws.cell(row=start_row + i, column=2, value=preco_formatado)
        cell_preco.number_format = 'R$ #,##0.00'
        
        # Link
        cell_link = ws.cell(row=start_row + i, column=3, value="Ver Oferta")
        if linha[2]:
            cell_link.hyperlink = linha[2]
            cell_link.style = "Hyperlink"

    # Aplica formatação
    formatar_planilha(ws, loja)

    # Tenta salvar
    try:
        wb.save(ARQUIVO_EXCEL)
        print(f"Dados salvos na aba '{loja}' com sucesso!")
    except PermissionError:
        print(f"\n[ERRO CRÍTICO] Não foi possível salvar '{ARQUIVO_EXCEL}'. Arquivo aberto?")

"""Abrir arquivo exel"""
def abrir_arquivo_excel():
    """Tenta abrir o arquivo Excel no sistema operacional."""
    if not os.path.exists(ARQUIVO_EXCEL):
        print("Arquivo ainda não existe.")
        return
    print(f"Abrindo {ARQUIVO_EXCEL}...")
    os.startfile(ARQUIVO_EXCEL)

