from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def mostrar_janela_exel(master=None):
    """Mostra uma janela simples informando que a função está em desenvolvimento."""
    
    janela = tk.Toplevel(master)
    janela.title("Interagir com Excel")
    janela.geometry("600x500")
    janela.configure(bg="#f0f2f5")
    janela.resizable(False, False)

    # Commando pra fixar a janela na frente
    janela.attributes('-topmost', True)


    fonte_geral = ("Segoe UI", 10)

    lbl_info = tk.Label(
        janela,
        text="Função de interagir com Excel\nestá em desenvolvimento....",
        font=fonte_geral,
        pady=20
    )
    lbl_info.pack()

    # botao para somar valores do excel
    bnt_verificar = tk.Button(
        janela,
        text="Somar Todos os Valores do Excel",
        font=fonte_geral,
        command=somar_valores_excel
    )
    bnt_verificar.pack(pady=10)


    # botao para retornar o menor valor do excel
    bnt_menor = tk.Button(
        janela,
        text="Retornar o Menor Valor do Excel",
        font=fonte_geral,
        command=menor_valor_excel
    )
    bnt_menor.pack(pady=10)


    #botão paraabrir o arquivo excel
    btn_abrir_excel = tk.Button(
        janela,
        text="Abrir Arquivo Excel",
        font=fonte_geral,
        command=abrir_arquivo_excel
    )
    btn_abrir_excel.pack(pady=10)


    # Botão para fechar a janela
    btn_fechar = tk.Button(
        janela,
        text="Fechar",
        command=janela.destroy,
        font=fonte_geral
    )

    btn_fechar.pack(side="bottom", pady=20)


#verifica se o arquivo excel existe
def verificar_arquivo_excel():
    """Verifica se o arquivo Excel existe """

    # Nome do arquivo Excel a ser verificado
    arquivo = "comparativo_precos.xlsx"

    try:
        # Tenta abrir apenas para ver se existe e é válido
        workbook = load_workbook(arquivo)
        workbook.close()
        print("Arquivo Encontrado e válido.")
        return True
    
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo não encontrado:\n{arquivo}")
        return False
    
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao abrir o arquivo:\n{e}")
        return False
    

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from tkinter import messagebox


def somar_valores_excel():
    """Soma apenas os valores da coluna de Preço (coluna B)."""

    if not verificar_arquivo_excel():
        return

    arquivo = "comparativo_precos.xlsx"

    try:
        workbook = load_workbook(arquivo)
        sheet = workbook.active

        soma_total = 0.0

        # Começa da linha 2 (pulando cabeçalho)
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            cell = row[0]

            if isinstance(cell, (int, float)):
                soma_total += cell
            elif isinstance(cell, str):
                soma_total += tratar_preco(cell)

        # Estilo
        alinhar_centro = Alignment(horizontal="center", vertical="center")

        # Ajuste de coluna
        sheet.column_dimensions['E'].width = 25

        # Escreve resultado
        sheet["E3"] = "Soma Total dos Valores:"
        sheet["E4"] = soma_total

        sheet["E3"].alignment = alinhar_centro
        sheet["E4"].alignment = alinhar_centro

        # Formatação monetária no Excel
        sheet["E4"].number_format = 'R$ #,##0.00'

        # Salva antes de fechar
        workbook.save(arquivo)
        workbook.close()

        messagebox.showinfo("Resultado", f"A soma total dos preços é: R$ {soma_total:,.2f}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")





def menor_valor_excel():
    """Escreve o menor valor da coluna de Preço ao lado da Soma Total."""

    if not verificar_arquivo_excel():
        return

    arquivo = "comparativo_precos.xlsx"

    try:
        workbook = load_workbook(arquivo)
        sheet = workbook.active

        menor_valor = None

        # Percorre somente a coluna B (Preço), ignorando cabeçalho
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            cell = row[0]

            if isinstance(cell, (int, float)):
                valor = float(cell)
            elif isinstance(cell, str):
                valor = tratar_preco(cell)
            else:
                continue

            if menor_valor is None or valor < menor_valor:
                menor_valor = valor

        if menor_valor is None:
            messagebox.showinfo("Resultado", "Nenhum valor numérico encontrado.")
            workbook.close()
            return

        alinhar_centro = Alignment(horizontal="center", vertical="center")

        # Ajusta largura da coluna G
        sheet.column_dimensions['G'].width = 20

        # Escreve ao lado da soma
        sheet["G3"] = "Menor Valor:"
        sheet["G4"] = menor_valor

        sheet["G3"].alignment = alinhar_centro
        sheet["G4"].alignment = alinhar_centro

        # Formato moeda
        sheet["G4"].number_format = 'R$ #,##0.00'

        workbook.save(arquivo)
        workbook.close()

        messagebox.showinfo("Resultado", f"O menor preço é: R$ {menor_valor:,.2f}")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {e}")




def tratar_preco(valor):
    """Converte valor monetário BR/US para float corretamente."""
    
    if not valor:
        return 0.0

    if isinstance(valor, (int, float)):
        return float(valor)

    valor_str = str(valor).strip().lower()
    
    # Remove símbolos
    valor_str = valor_str.replace('r$', '').replace('us$', '').strip()

    if ',' in valor_str:
        valor_str = valor_str.replace('.', '')  # remove milhar
        valor_str = valor_str.replace(',', '.')  # vírgula vira ponto

    elif '.' in valor_str:
        partes = valor_str.split('.')
        
        if len(partes[-1]) == 3:
            valor_str = valor_str.replace('.', '')
    
    try:
        return float(valor_str)
    except ValueError:
        return 0.0


#função abrir arquivo excel
def abrir_arquivo_excel():
    """Abre o arquivo Excel usando o aplicativo padrão do sistema."""
    try:
        os.startfile("comparativo_precos.xlsx")
    except:
        messagebox.showerror("Erro", "Não foi possível abrir o arquivo Excel verifique se ele existe.")
